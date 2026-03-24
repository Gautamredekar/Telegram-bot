import os
import re
import json
import time
import random
import string
import threading
from collections import Counter
from openpyxl import load_workbook
from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup
from telegram.ext import Updater, CommandHandler, CallbackQueryHandler, CallbackContext, MessageHandler, Filters, ConversationHandler

TOKEN = "8662131088:AAF797XeQt_29ebyvTuSgFN_zmxe2W7mtII"
ADMIN_ID = 6808834320
ADMIN_USERNAME = "ERWIN_HEREE"
DB_FILE = "database.json"

def load_db():
    try:
        with open(DB_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
    except:
        data = {}
    if "users" not in data:
        data["users"] = {}
    if "keys" not in data:
        data["keys"] = {}
    if "plans" not in data:
        data["plans"] = {
            "week": {"price": 50, "duration": 604800},
            "month": {"price": 150, "duration": 2592000},
            "3month": {"price": 300, "duration": 7776000}
        }
    return data

def save_db(data):
    with open(DB_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4)

def generate_key():
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=10))

def get_expiry(duration):
    return int(time.time()) + duration

def is_admin(user_id):
    return int(user_id) == ADMIN_ID

def has_access(user_id):
    if is_admin(user_id):
        return True
    db = load_db()
    user = db["users"].get(str(user_id))
    if not user:
        return False
    if user.get("blocked", False):
        return False
    return int(user.get("expiry", 0)) > int(time.time())

def require_access(update):
    if has_access(update.effective_user.id):
        return True
    update.message.reply_text("❌ You don't have access. Contact admin to get key.")
    return False

def tools_keyboard():
    return ReplyKeyboardMarkup(
        [
            ["Split File", "Combine Files"],
            ["Rename VCF", "TXT → VCF"],
            ["VCF → TXT", "XLSX → TXT"],
            ["Numbers → TXT/VCF"],
        ],
        resize_keyboard=True
    )

def show_tools(message):
    message.reply_text("Choose a tool:", reply_markup=tools_keyboard())

def start(update: Update, context: CallbackContext):
    user_id = str(update.effective_user.id)
    db = load_db()

    if user_id not in db["users"]:
        db["users"][user_id] = {
            "expiry": get_expiry(3600),
            "blocked": False,
            "notified": False
        }
        save_db(db)
        update.message.reply_text("🎉 Free Demo Activated (1 Hour)")

    keyboard = [
        [InlineKeyboardButton("🔑 Redeem Key", callback_data="redeem")],
        [InlineKeyboardButton("💰 Buy Key", callback_data="buy")],
        [InlineKeyboardButton("🧰 Open Tools", callback_data="open_tools")],
        [InlineKeyboardButton("📞 Contact Admin", url=f"https://t.me/{ADMIN_USERNAME}")]
    ]

    if is_admin(update.effective_user.id):
        keyboard.append([InlineKeyboardButton("⚙️ Admin Panel", callback_data="admin_panel")])

    update.message.reply_text("🤖 Welcome", reply_markup=InlineKeyboardMarkup(keyboard))

    if has_access(update.effective_user.id):
        show_tools(update.message)

def buttons(update: Update, context: CallbackContext):
    query = update.callback_query
    query.answer()
    db = load_db()

    try:
        if query.data == "buy":
            msg = "💰 Plans:\n\n"
            for name, data in db["plans"].items():
                msg += f"{name} - ₹{data['price']}\n"
            msg += f"\nContact: @{ADMIN_USERNAME}"
            query.edit_message_text(msg)

        elif query.data == "redeem":
            query.edit_message_text("Send:\n/redeem KEY")

        elif query.data == "open_tools":
            if not has_access(query.from_user.id):
                query.edit_message_text("❌ Access expired / not active")
                return
            query.message.reply_text("Choose a tool:", reply_markup=tools_keyboard())

        elif query.data == "admin_panel":
            if not is_admin(query.from_user.id):
                return

            keyboard = [
                [InlineKeyboardButton("🔑 Generate Key", callback_data="gen")],
                [InlineKeyboardButton("💰 Set Price", callback_data="set_price")],
                [InlineKeyboardButton("➕ Add Plan", callback_data="add_plan")],
                [InlineKeyboardButton("👥 Users", callback_data="show_users")],
                [InlineKeyboardButton("📢 Broadcast", callback_data="broadcast_msg")],
                [InlineKeyboardButton("⛔ Block User", callback_data="block_info")],
                [InlineKeyboardButton("✅ Unblock User", callback_data="unblock_info")],
            ]
            query.edit_message_text("⚙️ ADMIN PANEL", reply_markup=InlineKeyboardMarkup(keyboard))

        elif query.data == "gen":
            keyboard = []
            for plan in db["plans"]:
                keyboard.append([InlineKeyboardButton(plan, callback_data=f"gen_{plan}")])

            query.edit_message_text(
                "Select plan to generate key:",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )

        elif query.data.startswith("gen_"):
            plan = query.data.split("_", 1)[1]
            if plan not in db["plans"]:
                return query.edit_message_text("❌ Plan not found")

            key = generate_key()
            db["keys"][key] = {"plan": plan, "used": False}
            save_db(db)

            query.edit_message_text(f"✅ Key Generated\n\n🔑 {key}\n📦 Plan: {plan}")

        elif query.data == "set_price":
            query.edit_message_text("Use:\n/setprice plan price")

        elif query.data == "add_plan":
            query.edit_message_text("Use:\n/addplan name duration_seconds price")

        elif query.data == "show_users":
            users_list = list(db["users"].keys())
            if not users_list:
                return query.edit_message_text("No users found")

            msg = "👥 Users:\n\n"
            for uid in users_list:
                msg += f"{uid}\n"

            if len(msg) > 4000:
                for i in range(0, len(msg), 4000):
                    context.bot.send_message(query.from_user.id, msg[i:i+4000])
                query.answer("Sent in parts")
            else:
                query.edit_message_text(msg)

        elif query.data == "broadcast_msg":
            query.edit_message_text("Use:\n/broadcast your message")

        elif query.data == "block_info":
            query.edit_message_text("Use:\n/blockuser user_id")

        elif query.data == "unblock_info":
            query.edit_message_text("Use:\n/unblockuser user_id")

    except Exception as e:
        query.message.reply_text(f"Error: {e}")

def redeem(update: Update, context: CallbackContext):
    user_id = str(update.effective_user.id)

    if len(context.args) != 1:
        return update.message.reply_text("Usage: /redeem KEY")

    key = context.args[0]
    db = load_db()

    if key not in db["keys"]:
        return update.message.reply_text("Invalid key")

    if db["keys"][key]["used"]:
        return update.message.reply_text("Already used")

    plan = db["keys"][key]["plan"]
    duration = db["plans"][plan]["duration"]

    db["users"][user_id] = {
        "expiry": get_expiry(duration),
        "blocked": False,
        "notified": False
    }

    db["keys"][key]["used"] = True
    save_db(db)

    update.message.reply_text("✅ Activated")
    show_tools(update.message)

def admin(update: Update, context: CallbackContext):
    if not is_admin(update.effective_user.id):
        return

    keyboard = [
        [InlineKeyboardButton("🔑 Generate Key", callback_data="gen")],
        [InlineKeyboardButton("💰 Set Price", callback_data="set_price")],
        [InlineKeyboardButton("➕ Add Plan", callback_data="add_plan")],
        [InlineKeyboardButton("👥 Users", callback_data="show_users")],
        [InlineKeyboardButton("📢 Broadcast", callback_data="broadcast_msg")],
        [InlineKeyboardButton("⛔ Block User", callback_data="block_info")],
        [InlineKeyboardButton("✅ Unblock User", callback_data="unblock_info")],
    ]

    update.message.reply_text("⚙️ ADMIN PANEL", reply_markup=InlineKeyboardMarkup(keyboard))

def genkey(update: Update, context: CallbackContext):
    if not is_admin(update.effective_user.id):
        return
    if len(context.args) != 1:
        return update.message.reply_text("Usage: /genkey plan")
    plan = context.args[0]
    db = load_db()
    if plan not in db["plans"]:
        return update.message.reply_text("Plan not found")
    key = generate_key()
    db["keys"][key] = {"plan": plan, "used": False}
    save_db(db)
    update.message.reply_text(f"Key: {key}")

def setprice(update: Update, context: CallbackContext):
    if not is_admin(update.effective_user.id):
        return
    if len(context.args) != 2:
        return update.message.reply_text("Usage: /setprice plan price")
    plan = context.args[0]
    try:
        price = int(context.args[1])
    except:
        return update.message.reply_text("Price must be number")
    db = load_db()
    if plan not in db["plans"]:
        return update.message.reply_text("Plan not found")
    db["plans"][plan]["price"] = price
    save_db(db)
    update.message.reply_text("Price updated")

def addplan(update: Update, context: CallbackContext):
    if not is_admin(update.effective_user.id):
        return
    if len(context.args) != 3:
        return update.message.reply_text("Usage: /addplan name duration_seconds price")
    name = context.args[0]
    try:
        duration = int(context.args[1]); price = int(context.args[2])
    except:
        return update.message.reply_text("Duration & price must be numbers")
    db = load_db()
    db["plans"][name] = {"duration": duration, "price": price}
    save_db(db)
    update.message.reply_text("Plan added")

def users(update: Update, context: CallbackContext):
    if not is_admin(update.effective_user.id):
        return
    db = load_db()
    msg = "👥 Users:\n\n"
    for uid in db["users"]:
        msg += f"{uid}\n"
    update.message.reply_text(msg[:4000] if msg else "No users")

def broadcast(update: Update, context: CallbackContext):
    if not is_admin(update.effective_user.id):
        return
    if not context.args:
        return update.message.reply_text("Usage: /broadcast message")
    msg = " ".join(context.args)
    db = load_db()
    success = 0
    fail = 0
    for uid in db["users"]:
        try:
            context.bot.send_message(int(uid), msg)
            success += 1
        except:
            fail += 1
    update.message.reply_text(f"Sent: {success} | Failed: {fail}")

def blockuser(update: Update, context: CallbackContext):
    if not is_admin(update.effective_user.id):
        return
    if len(context.args) != 1:
        return update.message.reply_text("Usage: /blockuser user_id")
    uid = str(context.args[0])
    db = load_db()
    if uid not in db["users"]:
        db["users"][uid] = {"expiry": 0, "blocked": True, "notified": False}
    else:
        db["users"][uid]["blocked"] = True
    save_db(db)
    update.message.reply_text("User blocked")

def unblockuser(update: Update, context: CallbackContext):
    if not is_admin(update.effective_user.id):
        return
    if len(context.args) != 1:
        return update.message.reply_text("Usage: /unblockuser user_id")
    uid = str(context.args[0])
    db = load_db()
    if uid in db["users"]:
        db["users"][uid]["blocked"] = False
        save_db(db)
        update.message.reply_text("User unblocked")
    else:
        update.message.reply_text("User not found")

def expiry_checker(bot):
    while True:
        db = load_db()
        now = time.time()
        for uid, data in db["users"].items():
            if data["expiry"] < now and not data.get("notified"):
                try:
                    bot.send_message(int(uid), "⏳ Plan expired!")
                    db["users"][uid]["notified"] = True
                except:
                    pass
        save_db(db)
        time.sleep(60)

def format_number(num):
    num = num.strip().replace(" ", "").replace("-", "")
    if num.startswith("+"):
        num = num[1:]
    return num

def detect_country_code(numbers):
    prefixes = []
    for num in numbers:
        if len(num) >= 3:
            prefixes.append(num[:2])
            prefixes.append(num[:3])
    return Counter(prefixes).most_common(1)[0][0]

# states
TXT2VCF_FILE, TXT2VCF_NAME, TXT2VCF_CONTACT, TXT2VCF_LIMIT = range(100, 104)
RENAME_ASK_FILES, RENAME_ASK_CONTACT_NAME, RENAME_ASK_FILE_NAME = range(110, 113)
COMBINE_ASK_FILES, COMBINE_ASK_NAME = range(120, 122)
NUM_ASK_NUMBERS, NUM_ASK_TYPE, NUM_ASK_TXT_NAME, NUM_ASK_VCF_NAME, NUM_ASK_CONTACT_NAME, NUM_ASK_PER_FILE = range(130, 136)
VCF2TXT_FILES, VCF2TXT_NAME = range(140, 142)
XLSX_ASK_FILE, XLSX_ASK_FILENAME = range(150, 152)
SPLIT_ASK_MODE, SPLIT_ASK_FILE, SPLIT_ASK_PER_FILE, SPLIT_ASK_TOTAL_FILES, SPLIT_ASK_FILENAME, SPLIT_ASK_CONTACT_NAME = range(160, 166)

user_data_store = {}

def txt2vcf_start(update: Update, context: CallbackContext):
    if not require_access(update):
        return ConversationHandler.END
    update.message.reply_text("📂 Send TXT file to start")
    return TXT2VCF_FILE

def txt2vcf_handle_document(update: Update, context: CallbackContext):
    file = update.message.document
    user_id = update.message.chat_id
    if not file.file_name.endswith(".txt"):
        update.message.reply_text("❌ Send only TXT file")
        return TXT2VCF_FILE
    file_path = f"{user_id}.txt"
    file.get_file().download(file_path)
    with open(file_path, "r") as f:
        numbers = [line.strip() for line in f if line.strip()]
    if not numbers:
        update.message.reply_text("❌ TXT file is empty")
        return TXT2VCF_FILE
    user_data_store[user_id] = {"numbers": numbers, "source_file": file_path}
    update.message.reply_text(f"✅ Total Numbers: {len(numbers)}\n\nEnter VCF file name (Example: Erwin)")
    return TXT2VCF_NAME

def txt2vcf_handle_filename(update: Update, context: CallbackContext):
    user_id = update.message.chat_id
    user_data_store[user_id]["filename"] = update.message.text.strip()
    update.message.reply_text("Enter contact name (Example: Erwin)")
    return TXT2VCF_CONTACT

def txt2vcf_handle_contact(update: Update, context: CallbackContext):
    user_id = update.message.chat_id
    user_data_store[user_id]["contact"] = update.message.text.strip()
    update.message.reply_text("How many contacts per VCF?")
    return TXT2VCF_LIMIT

def txt2vcf_generate(update: Update, context: CallbackContext):
    user_id = update.message.chat_id
    text = update.message.text.strip()
    try:
        limit = int(text)
        if limit <= 0:
            raise ValueError
        user_data_store[user_id]["limit"] = limit
    except:
        update.message.reply_text("❌ Enter valid number")
        return TXT2VCF_LIMIT
    data = user_data_store.get(user_id)
    if not data:
        update.message.reply_text("❌ Error, send TXT again")
        return ConversationHandler.END
    numbers = data["numbers"]
    filename = data["filename"]
    contact_name = data["contact"]
    limit = data["limit"]
    file_index = 1
    contact_index = 1
    current_count = 0
    vcf_name = f"{filename}.vcf"
    vcf = open(vcf_name, "w")
    update.message.reply_text("⚡ Processing...")
    for num in numbers:
        num = format_number(num)
        if not num:
            continue
        name = f"{contact_name} {contact_index}"
        vcf.write("BEGIN:VCARD\n")
        vcf.write("VERSION:3.0\n")
        vcf.write(f"FN:{name}\n")
        vcf.write(f"TEL:+{num}\n")
        vcf.write("END:VCARD\n\n")
        contact_index += 1
        current_count += 1
        if current_count == limit:
            vcf.close()
            if os.path.exists(vcf_name) and os.path.getsize(vcf_name) > 0:
                context.bot.send_document(chat_id=user_id, document=open(vcf_name, "rb"))
            os.remove(vcf_name)
            file_index += 1
            current_count = 0
            vcf_name = f"{filename} {file_index}.vcf"
            vcf = open(vcf_name, "w")
    vcf.close()
    if os.path.exists(vcf_name) and os.path.getsize(vcf_name) > 0:
        context.bot.send_document(chat_id=user_id, document=open(vcf_name, "rb"))
        os.remove(vcf_name)
    source_file = data.get("source_file")
    if source_file and os.path.exists(source_file):
        os.remove(source_file)
    update.message.reply_text("✅ Done")
    return ConversationHandler.END

def rename_start(update: Update, context: CallbackContext):
    if not require_access(update):
        return ConversationHandler.END
    user_data_store[update.effective_user.id] = {"files": []}
    update.message.reply_text("Send one or multiple .vcf files:")
    return RENAME_ASK_FILES

def rename_get_files(update: Update, context: CallbackContext):
    file = update.message.document
    if not file.file_name.endswith(".vcf"):
        update.message.reply_text("Send only .vcf files.")
        return RENAME_ASK_FILES
    file_path = file.get_file().download()
    user_data_store[update.effective_user.id]["files"].append(file_path)
    update.message.reply_text("File received.\nSend more or type 'done' when finished.")
    return RENAME_ASK_FILES

def rename_done_files(update: Update, context: CallbackContext):
    if not user_data_store[update.effective_user.id]["files"]:
        update.message.reply_text("No files received. Send at least one .vcf file.")
        return RENAME_ASK_FILES
    update.message.reply_text("Enter contact name (Example: Erwin):")
    return RENAME_ASK_CONTACT_NAME

def rename_ask_file_name(update: Update, context: CallbackContext):
    contact_name = update.message.text.strip()
    user_data_store[update.effective_user.id]["contact_name"] = contact_name
    update.message.reply_text("Enter VCF file name (Example: Erwin):")
    return RENAME_ASK_FILE_NAME

def rename_process_vcf(update: Update, context: CallbackContext):
    data = user_data_store[update.effective_user.id]
    files = data["files"]
    contact_name = data["contact_name"]
    base_name = update.message.text.strip()
    contact_count = 1
    file_index = 1
    for file_path in files:
        with open(file_path, "r") as f:
            content = f.read()
        entries = content.split("END:VCARD")
        entries = [e.strip() for e in entries if e.strip()]
        output_name = f"{base_name} {file_index}.vcf"
        with open(output_name, "w") as f:
            for entry in entries:
                lines = entry.splitlines()
                number_line = [l for l in lines if l.startswith("TEL")]
                number = number_line[0].split(":")[1] if number_line else ""
                f.write("BEGIN:VCARD\n")
                f.write("VERSION:3.0\n")
                f.write(f"FN:{contact_name} {contact_count}\n")
                f.write(f"TEL:{number}\n")
                f.write("END:VCARD\n")
                contact_count += 1
        update.message.reply_document(open(output_name, "rb"))
        os.remove(output_name)
        os.remove(file_path)
        file_index += 1
    update.message.reply_text("✅ VCF renaming completed!")
    return ConversationHandler.END

def combine_start(update: Update, context: CallbackContext):
    if not require_access(update):
        return ConversationHandler.END
    user_data_store[update.effective_user.id] = {"files": [], "type": None}
    update.message.reply_text("Send multiple .vcf or .txt files.\nType 'done' when finished.")
    return COMBINE_ASK_FILES

def combine_get_files(update: Update, context: CallbackContext):
    file = update.message.document
    file_name = file.file_name
    if not file_name.endswith((".vcf", ".txt")):
        update.message.reply_text("Only .vcf or .txt files allowed.")
        return COMBINE_ASK_FILES
    file_type = ".vcf" if file_name.endswith(".vcf") else ".txt"
    if user_data_store[update.effective_user.id]["type"] is None:
        user_data_store[update.effective_user.id]["type"] = file_type
    elif user_data_store[update.effective_user.id]["type"] != file_type:
        update.message.reply_text("❌ Send same file type only (all .vcf or all .txt).")
        return COMBINE_ASK_FILES
    file_path = file.get_file().download()
    user_data_store[update.effective_user.id]["files"].append(file_path)
    update.message.reply_text("File added. Send more or type 'done'.")
    return COMBINE_ASK_FILES

def combine_ask_name(update: Update, context: CallbackContext):
    data = user_data_store[update.effective_user.id]
    if not data["files"]:
        update.message.reply_text("No files received.")
        return COMBINE_ASK_FILES
    update.message.reply_text("Enter output file name (Example: Erwin):")
    return COMBINE_ASK_NAME

def combine_files(update: Update, context: CallbackContext):
    data = user_data_store[update.effective_user.id]
    base_name = update.message.text.strip()
    files = data["files"]
    file_type = data["type"]
    output_name = f"{base_name}{file_type}"
    with open(output_name, "w") as outfile:
        for file_path in files:
            with open(file_path, "r") as infile:
                content = infile.read().strip()
                if content:
                    outfile.write(content + "\n")
            os.remove(file_path)
    update.message.reply_document(open(output_name, "rb"))
    os.remove(output_name)
    update.message.reply_text("✅ Files combined successfully!")
    return ConversationHandler.END

def num_start(update: Update, context: CallbackContext):
    if not require_access(update):
        return ConversationHandler.END
    update.message.reply_text("Send numbers (one per line):")
    return NUM_ASK_NUMBERS

def num_get_numbers(update: Update, context: CallbackContext):
    lines = update.message.text.strip().split("\n")
    raw_numbers = []
    for num in lines:
        num = num.strip().replace(" ", "")
        if num.isdigit():
            raw_numbers.append(num)
    if not raw_numbers:
        update.message.reply_text("Send valid numbers only.")
        return NUM_ASK_NUMBERS
    country_code = detect_country_code(raw_numbers)
    clean_numbers = []
    for num in raw_numbers:
        if not num.startswith(country_code):
            num = country_code + num
        clean_numbers.append(num)
    user_data_store[update.effective_user.id] = {
        "numbers": clean_numbers,
        "country_code": country_code
    }
    update.message.reply_text(f"Detected country code: +{country_code}\n\nChoose: .txt or .vcf")
    return NUM_ASK_TYPE

def num_ask_type(update: Update, context: CallbackContext):
    text = update.message.text.lower().strip()
    if text == ".txt":
        update.message.reply_text("Enter TXT file name:")
        return NUM_ASK_TXT_NAME
    elif text == ".vcf":
        update.message.reply_text("Enter VCF file base name (Example: Erwin):")
        return NUM_ASK_VCF_NAME
    else:
        update.message.reply_text("Type .txt or .vcf")
        return NUM_ASK_TYPE

def num_create_txt(update: Update, context: CallbackContext):
    file_name = update.message.text.strip() + ".txt"
    numbers = user_data_store[update.effective_user.id]["numbers"]
    with open(file_name, "w") as f:
        for num in numbers:
            f.write(num + "\n")
    update.message.reply_document(open(file_name, "rb"))
    os.remove(file_name)
    return ConversationHandler.END

def num_ask_contact_name(update: Update, context: CallbackContext):
    user_data_store[update.effective_user.id]["vcf_name"] = update.message.text.strip()
    update.message.reply_text("Enter contact name (Example: Erwin):")
    return NUM_ASK_CONTACT_NAME

def num_ask_per_file(update: Update, context: CallbackContext):
    user_data_store[update.effective_user.id]["contact_name"] = update.message.text.strip()
    update.message.reply_text("How many contacts per VCF file?")
    return NUM_ASK_PER_FILE

def num_create_vcf(update: Update, context: CallbackContext):
    try:
        per_file = int(update.message.text.strip())
    except:
        update.message.reply_text("Enter a valid number.")
        return NUM_ASK_PER_FILE
    data = user_data_store[update.effective_user.id]
    numbers = data["numbers"]
    base_name = data["vcf_name"]
    contact_name = data["contact_name"]
    files = []
    chunk = []
    count = 1
    file_index = 1
    for num in numbers:
        chunk.append(num)
        if len(chunk) == per_file:
            filename = f"{base_name}.vcf" if file_index == 1 else f"{base_name} {file_index}.vcf"
            with open(filename, "w") as f:
                for number in chunk:
                    f.write("BEGIN:VCARD\n")
                    f.write("VERSION:3.0\n")
                    f.write(f"FN:{contact_name} {count}\n")
                    f.write(f"TEL:+{number}\n")
                    f.write("END:VCARD\n")
                    count += 1
            files.append(filename)
            chunk = []
            file_index += 1
    if chunk:
        filename = f"{base_name} {file_index}.vcf"
        with open(filename, "w") as f:
            for number in chunk:
                f.write("BEGIN:VCARD\n")
                f.write("VERSION:3.0\n")
                f.write(f"FN:{contact_name} {count}\n")
                f.write(f"TEL:+{number}\n")
                f.write("END:VCARD\n")
                count += 1
        files.append(filename)
    for file in files:
        update.message.reply_document(open(file, "rb"))
        os.remove(file)
    return ConversationHandler.END

def extract_numbers(vcf_path):
    numbers = []
    with open(vcf_path, "r", errors="ignore") as f:
        for line in f:
            if "TEL" in line:
                num = re.sub(r"[^\d+]", "", line)
                if num:
                    if num.startswith("+"):
                        num = num[1:]
                    numbers.append(num)
    return numbers

def vcf2txt_start(update: Update, context: CallbackContext):
    if not require_access(update):
        return ConversationHandler.END
    user_data_store[update.message.chat_id] = {"files": []}
    update.message.reply_text("📂 Send multiple VCF files\n\nWhen done, type /done")
    return VCF2TXT_FILES

def vcf2txt_handle_document(update: Update, context: CallbackContext):
    file = update.message.document
    user_id = update.message.chat_id
    if not file.file_name.endswith(".vcf"):
        update.message.reply_text("❌ Send only VCF files")
        return VCF2TXT_FILES
    file_path = f"{user_id}_{file.file_name}"
    file.get_file().download(file_path)
    if user_id not in user_data_store:
        user_data_store[user_id] = {"files": []}
    user_data_store[user_id]["files"].append(file_path)
    update.message.reply_text(f"✅ Added: {file.file_name}")
    return VCF2TXT_FILES

def vcf2txt_done(update: Update, context: CallbackContext):
    user_id = update.message.chat_id
    if user_id not in user_data_store or not user_data_store[user_id]["files"]:
        update.message.reply_text("❌ No VCF files uploaded")
        return VCF2TXT_FILES
    update.message.reply_text("Enter output TXT file name (Example: contacts)")
    return VCF2TXT_NAME

def vcf2txt_generate(update: Update, context: CallbackContext):
    user_id = update.message.chat_id
    filename = update.message.text.strip()
    files = user_data_store[user_id]["files"]
    all_numbers = []
    update.message.reply_text("⚡ Processing...")
    for file_path in files:
        nums = extract_numbers(file_path)
        all_numbers.extend(nums)
        os.remove(file_path)
    if not all_numbers:
        update.message.reply_text("❌ No numbers found")
        return ConversationHandler.END
    txt_name = f"{filename}.txt"
    with open(txt_name, "w") as f:
        for num in all_numbers:
            f.write(num + "\n")
    context.bot.send_document(chat_id=user_id, document=open(txt_name, "rb"))
    os.remove(txt_name)
    user_data_store[user_id] = {"files": []}
    update.message.reply_text(f"✅ Done\nTotal Numbers: {len(all_numbers)}")
    return ConversationHandler.END

def xlsx_start(update: Update, context: CallbackContext):
    if not require_access(update):
        return ConversationHandler.END
    update.message.reply_text(
        "📂 Send .xlsx file\n"
        "I will convert it to .txt (1 line = 1 value)"
    )
    return XLSX_ASK_FILE

def xlsx_handle_file(update: Update, context: CallbackContext):
    doc = update.message.document
    if not doc.file_name.endswith(".xlsx"):
        update.message.reply_text("❌ Send only .xlsx file")
        return XLSX_ASK_FILE
    file = doc.get_file()
    file_path = doc.file_name
    file.download(file_path)
    context.user_data['file_path'] = file_path
    update.message.reply_text("✏️ Send output file name (without .txt)")
    return XLSX_ASK_FILENAME

def xlsx_convert_file(update: Update, context: CallbackContext):
    filename = update.message.text.strip()
    input_file = context.user_data.get("file_path")
    try:
        wb = load_workbook(input_file)
        sheet = wb.active
        output_file = f"{filename}.txt"
        with open(output_file, "w", encoding="utf-8") as f:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        f.write(str(cell).strip() + "\n")
        update.message.reply_document(open(output_file, "rb"))
        os.remove(input_file)
        os.remove(output_file)
    except Exception as e:
        update.message.reply_text(f"❌ Error: {e}")
    return ConversationHandler.END

def xlsx_cancel(update: Update, context: CallbackContext):
    update.message.reply_text("❌ Cancelled")
    return ConversationHandler.END

def split_start(update: Update, context: CallbackContext):
    if not require_access(update):
        return ConversationHandler.END
    update.message.reply_text("Type 'split' to start splitting files.")
    return SPLIT_ASK_MODE

def split_mode(update: Update, context: CallbackContext):
    if update.message.text.lower() == "split":
        update.message.reply_text("Send .txt or .vcf file:")
        return SPLIT_ASK_FILE
    else:
        update.message.reply_text("Type 'split' to continue.")
        return SPLIT_ASK_MODE

def split_get_file(update: Update, context: CallbackContext):
    file = update.message.document
    if not file.file_name.endswith((".txt", ".vcf")):
        update.message.reply_text("Send only .txt or .vcf file.")
        return SPLIT_ASK_FILE
    file_path = file.get_file().download()
    user_data_store[update.effective_user.id] = {"file": file_path}
    if file_path.endswith(".txt"):
        with open(file_path, "r") as f:
            entries = [l.strip() for l in f.readlines() if l.strip()]
    else:
        with open(file_path, "r") as f:
            content = f.read()
        raw = content.split("END:VCARD")
        entries = [e.strip() + "\nEND:VCARD\n" for e in raw if e.strip()]
    user_data_store[update.effective_user.id]["entries"] = entries
    total = len(entries)
    update.message.reply_text(f"Total numbers: {total}\n\nHow many numbers per file?")
    return SPLIT_ASK_PER_FILE

def split_ask_total_files(update: Update, context: CallbackContext):
    try:
        per_file = int(update.message.text)
    except:
        update.message.reply_text("Enter valid number.")
        return SPLIT_ASK_PER_FILE
    user_data_store[update.effective_user.id]["per_file"] = per_file
    update.message.reply_text("How many files to make?")
    return SPLIT_ASK_TOTAL_FILES

def split_ask_filename(update: Update, context: CallbackContext):
    try:
        total_files = int(update.message.text)
    except:
        update.message.reply_text("Enter valid number.")
        return SPLIT_ASK_TOTAL_FILES
    user_data_store[update.effective_user.id]["total_files"] = total_files
    update.message.reply_text("Enter output file name (Example: Erwin):")
    return SPLIT_ASK_FILENAME

def split_ask_contact_name(update: Update, context: CallbackContext):
    base_name = update.message.text.strip()
    user_data_store[update.effective_user.id]["base_name"] = base_name
    file_path = user_data_store[update.effective_user.id]["file"]
    if file_path.endswith(".vcf"):
        update.message.reply_text("Enter contact name (Example: Erwin):")
        return SPLIT_ASK_CONTACT_NAME
    else:
        return split_files(update, context)

def split_files(update: Update, context: CallbackContext):
    data = user_data_store[update.effective_user.id]
    entries = data["entries"]
    per_file = data["per_file"]
    total_files = data["total_files"]
    base_name = data["base_name"]
    file_path = data["file"]
    is_vcf = file_path.endswith(".vcf")
    contact_name = update.message.text.strip() if is_vcf else None
    index = 0
    contact_count = 1
    for i in range(total_files):
        chunk = entries[index:index + per_file]
        if not chunk:
            break
        filename = f"{base_name} {i+1}.vcf" if is_vcf else f"{base_name} {i+1}.txt"
        with open(filename, "w") as f:
            if is_vcf:
                for entry in chunk:
                    lines = entry.splitlines()
                    number_line = [l for l in lines if l.startswith("TEL")]
                    number = number_line[0].split(":")[1] if number_line else ""
                    f.write("BEGIN:VCARD\n")
                    f.write("VERSION:3.0\n")
                    f.write(f"FN:{contact_name} {contact_count}\n")
                    f.write(f"TEL:{number}\n")
                    f.write("END:VCARD\n")
                    contact_count += 1
            else:
                f.write("\n".join(chunk))
        update.message.reply_document(open(filename, "rb"))
        os.remove(filename)
        index += per_file
    os.remove(file_path)
    update.message.reply_text("✅ Splitting completed!")
    return ConversationHandler.END

def open_tools(update: Update, context: CallbackContext):
    if not require_access(update):
        return
    show_tools(update.message)

def main():
    updater = Updater(TOKEN, use_context=True)
    dp = updater.dispatcher
    dp.add_handler(CommandHandler("start", start))
    dp.add_handler(CommandHandler("tools", open_tools))
    dp.add_handler(CommandHandler("admin", admin))
    dp.add_handler(CommandHandler("redeem", redeem))
    dp.add_handler(CommandHandler("genkey", genkey))
    dp.add_handler(CommandHandler("setprice", setprice))
    dp.add_handler(CommandHandler("addplan", addplan))
    dp.add_handler(CommandHandler("users", users))
    dp.add_handler(CommandHandler("broadcast", broadcast))
    dp.add_handler(CommandHandler("blockuser", blockuser))
    dp.add_handler(CommandHandler("unblockuser", unblockuser))
    dp.add_handler(CallbackQueryHandler(buttons))

    dp.add_handler(ConversationHandler(
        entry_points=[CommandHandler("txt_to_vcf", txt2vcf_start), MessageHandler(Filters.regex("^TXT → VCF$"), txt2vcf_start)],
        states={
            TXT2VCF_FILE: [MessageHandler(Filters.document, txt2vcf_handle_document)],
            TXT2VCF_NAME: [MessageHandler(Filters.text & ~Filters.command, txt2vcf_handle_filename)],
            TXT2VCF_CONTACT: [MessageHandler(Filters.text & ~Filters.command, txt2vcf_handle_contact)],
            TXT2VCF_LIMIT: [MessageHandler(Filters.text & ~Filters.command, txt2vcf_generate)],
        },
        fallbacks=[CommandHandler("cancel", xlsx_cancel)],
    ))

    dp.add_handler(ConversationHandler(
        entry_points=[CommandHandler("rename_vcf", rename_start), MessageHandler(Filters.regex("^Rename VCF$"), rename_start)],
        states={
            RENAME_ASK_FILES: [
                MessageHandler(Filters.document, rename_get_files),
                MessageHandler(Filters.text & Filters.regex("^done$"), rename_done_files),
            ],
            RENAME_ASK_CONTACT_NAME: [MessageHandler(Filters.text & ~Filters.command, rename_ask_file_name)],
            RENAME_ASK_FILE_NAME: [MessageHandler(Filters.text & ~Filters.command, rename_process_vcf)],
        },
        fallbacks=[CommandHandler("cancel", xlsx_cancel)],
    ))

    dp.add_handler(ConversationHandler(
        entry_points=[CommandHandler("combine", combine_start), MessageHandler(Filters.regex("^Combine Files$"), combine_start)],
        states={
            COMBINE_ASK_FILES: [
                MessageHandler(Filters.document, combine_get_files),
                MessageHandler(Filters.text & Filters.regex("^done$"), combine_ask_name),
            ],
            COMBINE_ASK_NAME: [MessageHandler(Filters.text & ~Filters.command, combine_files)],
        },
        fallbacks=[CommandHandler("cancel", xlsx_cancel)],
    ))

    dp.add_handler(ConversationHandler(
        entry_points=[CommandHandler("numbers", num_start), MessageHandler(Filters.regex("^Numbers → TXT/VCF$"), num_start)],
        states={
            NUM_ASK_NUMBERS: [MessageHandler(Filters.text & ~Filters.command, num_get_numbers)],
            NUM_ASK_TYPE: [MessageHandler(Filters.text & ~Filters.command, num_ask_type)],
            NUM_ASK_TXT_NAME: [MessageHandler(Filters.text & ~Filters.command, num_create_txt)],
            NUM_ASK_VCF_NAME: [MessageHandler(Filters.text & ~Filters.command, num_ask_contact_name)],
            NUM_ASK_CONTACT_NAME: [MessageHandler(Filters.text & ~Filters.command, num_ask_per_file)],
            NUM_ASK_PER_FILE: [MessageHandler(Filters.text & ~Filters.command, num_create_vcf)],
        },
        fallbacks=[CommandHandler("cancel", xlsx_cancel)],
    ))

    dp.add_handler(ConversationHandler(
        entry_points=[CommandHandler("vcf_to_txt", vcf2txt_start), MessageHandler(Filters.regex("^VCF → TXT$"), vcf2txt_start)],
        states={
            VCF2TXT_FILES: [
                MessageHandler(Filters.document, vcf2txt_handle_document),
                CommandHandler("done", vcf2txt_done),
            ],
            VCF2TXT_NAME: [MessageHandler(Filters.text & ~Filters.command, vcf2txt_generate)],
        },
        fallbacks=[CommandHandler("cancel", xlsx_cancel)],
    ))

    dp.add_handler(ConversationHandler(
        entry_points=[CommandHandler("xlsx_to_txt", xlsx_start), MessageHandler(Filters.regex("^XLSX → TXT$"), xlsx_start)],
        states={
            XLSX_ASK_FILE: [MessageHandler(Filters.document, xlsx_handle_file)],
            XLSX_ASK_FILENAME: [MessageHandler(Filters.text & ~Filters.command, xlsx_convert_file)],
        },
        fallbacks=[CommandHandler("cancel", xlsx_cancel)],
    ))

    dp.add_handler(ConversationHandler(
        entry_points=[CommandHandler("split", split_start), MessageHandler(Filters.regex("^Split File$"), split_start)],
        states={
            SPLIT_ASK_MODE: [MessageHandler(Filters.text & ~Filters.command, split_mode)],
            SPLIT_ASK_FILE: [MessageHandler(Filters.document, split_get_file)],
            SPLIT_ASK_PER_FILE: [MessageHandler(Filters.text & ~Filters.command, split_ask_total_files)],
            SPLIT_ASK_TOTAL_FILES: [MessageHandler(Filters.text & ~Filters.command, split_ask_filename)],
            SPLIT_ASK_FILENAME: [MessageHandler(Filters.text & ~Filters.command, split_ask_contact_name)],
            SPLIT_ASK_CONTACT_NAME: [MessageHandler(Filters.text & ~Filters.command, split_files)],
        },
        fallbacks=[CommandHandler("cancel", xlsx_cancel)],
    ))

    threading.Thread(target=expiry_checker, args=(updater.bot,), daemon=True).start()
    print("🔥 BOT RUNNING...")
    updater.start_polling()
    updater.idle()

if __name__ == "__main__":
    main()
